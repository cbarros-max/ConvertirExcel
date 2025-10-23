import logging
import azure.functions as func
import pandas as pd
import io
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Python HTTP trigger function processed a request.")

    try:
        file = req.files.get('file')
        if not file:
            return func.HttpResponse("No se encontró el archivo en la solicitud. Se esperaba un campo 'file'.", status_code=400)

        file_content = file.stream.read()

        # Cargar el archivo Excel y omitir las primeras 6 filas
        df = pd.read_excel(
            io.BytesIO(file_content),
            engine='openpyxl' if file.filename.endswith('xlsx') else 'xlrd',
            skiprows=4
        )

        # Crear el Excel de salida
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Determinar el rango de la tabla correctamente
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            end_col_letter = get_column_letter(max_col)
            table_range = f"A1:{end_col_letter}{max_row}"

            # Crear y aplicar estilo de tabla
            table = Table(displayName="DataTable", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)

        output.seek(0)
        return func.HttpResponse(
            output.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename="converted.xlsx"',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            },
            status_code=200
        )

    except Exception as e:
        logging.exception("Error al procesar el archivo.")
        return func.HttpResponse(f"Error: {str(e)}", status_code=500)
